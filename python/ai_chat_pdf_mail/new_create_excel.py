import pandas as pd
import hashlib
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
import os

def calculate_checksum(file_path):
    hasher = hashlib.md5()
    with open(file_path, 'rb') as afile:
        buf = afile.read()
        hasher.update(buf)
    return hasher.hexdigest()

def create_excel_from_csv(input_csv):
    df = pd.read_csv(input_csv)

    total_amount = df['Invoice Amount'].sum()
    total_vat = df['VAT'].sum()
    totals_row = pd.DataFrame([{'Sender Name': 'Totals', 'Invoice Number': '', 'Invoice Amount': total_amount, 'VAT': total_vat}])

    description_text = os.getenv('sheet_description', 'Default Description')
    description_row = pd.DataFrame([{'Sender Name': description_text, 'Invoice Number': '', 'Invoice Amount': '', 'VAT': ''}])

    df = pd.concat([description_row, pd.DataFrame([{}]), df, pd.DataFrame([{}]), totals_row], ignore_index=True)

    time_for_filename = datetime.datetime.now().strftime("%H%M")
    checksum = calculate_checksum(input_csv)
    excel_file_name = f'invoices-{time_for_filename}-{checksum[:8]}.xlsx'
    df.to_excel(excel_file_name, index=False, header=False)

    workbook = openpyxl.load_workbook(excel_file_name)
    sheet = workbook.active

    # Set specific column widths
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 10

    # Alignment setup
    left_align = Alignment(horizontal='left')
    right_align = Alignment(horizontal='right')

    # Make the CSV header row (now the third row) bold and right-aligned
    bold_font = Font(bold=True)
    for cell in sheet["3:3"]:
        cell.font = bold_font
        cell.alignment = right_align

    # Left align the first two columns
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.alignment = left_align

    # Add CSV headers to the third row
    for i, column_name in enumerate(df.columns):
        sheet.cell(row=3, column=i+1, value=column_name)

    workbook.save(excel_file_name)

    return excel_file_name

# Example usage
excel_file_name

