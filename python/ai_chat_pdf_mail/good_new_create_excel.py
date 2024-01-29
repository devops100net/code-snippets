import pandas as pd
import hashlib
import datetime
import openpyxl
from openpyxl.styles import Font
import os

def calculate_checksum(file_path):
    hasher = hashlib.md5()
    with open(file_path, 'rb') as afile:
        buf = afile.read()
        hasher.update(buf)
    return hasher.hexdigest()

def create_excel_from_csv(input_csv):
    # Read data from CSV
    df = pd.read_csv(input_csv)

    # Totals row
    total_amount = df['Invoice Amount'].sum()
    total_vat = df['VAT'].sum()
    totals_row = pd.DataFrame([{'Sender Name': 'Totals', 'Invoice Number': '', 'Invoice Amount': total_amount, 'VAT': total_vat}])

    # Description row
    description_text = os.getenv('sheet_description', 'Default Description')
    description_row = pd.DataFrame([{'Sender Name': description_text, 'Invoice Number': '', 'Invoice Amount': '', 'VAT': ''}])

    # Combining rows
    df = pd.concat([description_row, pd.DataFrame([{}]), df, pd.DataFrame([{}]), totals_row], ignore_index=True)

    # Save to Excel file
    time_for_filename = datetime.datetime.now().strftime("%H%M")
    checksum = calculate_checksum(input_csv)
    excel_file_name = f'invoices-{time_for_filename}-{checksum[:8]}.xlsx'
    df.to_excel(excel_file_name, index=False, header=False)

    # Formatting in Excel
    workbook = openpyxl.load_workbook(excel_file_name)
    sheet = workbook.active
    
    # Set column widths and make the CSV header row (now the third row) bold
    for column in sheet.columns:
        sheet.column_dimensions[openpyxl.utils.get_column_letter(column[0].column)].width = 15
    for cell in sheet["3:3"]:
        cell.font = Font(bold=True)

    # Add CSV headers to the third row
    for i, column_name in enumerate(df.columns):
        sheet.cell(row=3, column=i+1, value=column_name)

    workbook.save(excel_file_name)

    return excel_file_name

# Example usage
excel_file_name = create_excel_from_csv('export-pdf.csv')

